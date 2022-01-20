# -- coding: UTF-8 --
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


def get_max_row(sheet):
    i = sheet.max_row
    real_max_row = 0
    while i > 0:
        row_dict = {i.value for i in sheet[i]}
        if row_dict == {None}:
            i = i-1
        else:
            real_max_row = i
            break
    return real_max_row + 1


def search_word(bb, word):
    browser = bb
    input_box = browser.find_element(by="id", value="text_to_transcribe")
    try:
        input_box.clear()
    except Exception as e:
        print("fail to clear")
    input_box.send_keys(word)
    browser.find_element(by="id", value="submit").click()
    element = browser.find_element(By.XPATH, "//span[contains(@class, 'transcribed_word')]").text
    return element


b = webdriver.Chrome()
b.get('https://tophonetics.com/zh/')
wb = load_workbook(r"./data/雅思词汇胜经词汇编辑.xlsx")
ws = wb.active
row_max = get_max_row(ws)
print(row_max)
for row in range(2, row_max):
    res = search_word(b, ws.cell(row, 2).value)
    ws.cell(row, 3).value = '/' + res + '/'
wb.save(r"./output/雅思词汇胜经词汇编辑.xlsx")
b.close()

